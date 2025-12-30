import argparse
import json
from pathlib import Path
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def load_data(ledger_path: str, bank_path: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    ledger = pd.read_csv(ledger_path, dtype=str, keep_default_na=False)
    bank = pd.read_csv(bank_path, dtype=str, keep_default_na=False)
    return ledger, bank


def standardize_columns(ledger_raw: pd.DataFrame, bank_raw: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    ledger = ledger_raw.copy()
    bank = bank_raw.copy()

    ledger = ledger.rename(
        columns={
            "txn_id": "ledger_id",
            "post_date": "ledger_date",
            "description": "ledger_desc",
            "amount": "ledger_amount_raw",
            "currency": "ledger_currency",
            "category": "ledger_category",
        }
    )

    bank = bank.rename(
        columns={
            "bank_id": "bank_id",
            "date": "bank_date",
            "details": "bank_desc",
            "amount_usd": "bank_amount_raw",
        }
    )

    for col in ["ledger_id", "ledger_date", "ledger_desc", "ledger_amount_raw"]:
        if col not in ledger.columns:
            raise ValueError(f"Missing required ledger column: {col}")

    for col in ["bank_id", "bank_date", "bank_desc", "bank_amount_raw"]:
        if col not in bank.columns:
            raise ValueError(f"Missing required bank column: {col}")

    return ledger, bank


def _clean_text(x: str) -> str:
    s = "" if x is None else str(x)
    return " ".join(s.strip().split())


def _norm_desc(x: str) -> str:
    return _clean_text(x).lower()


def _parse_date(x: str) -> pd.Timestamp:
    s = _clean_text(x)
    if s == "":
        return pd.NaT
    dt = pd.to_datetime(s, errors="coerce")
    return dt.normalize() if not pd.isna(dt) else pd.NaT


def _parse_amount_cents(x: str) -> int | None:
    s = _clean_text(x)
    if s == "":
        return None

    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1].strip()

    s = s.replace("$", "").replace(",", "").replace(" ", "")

    try:
        val = float(s)
    except ValueError:
        return None

    if neg:
        val = -val

    return int(round(val * 100))


def clean_fields(ledger: pd.DataFrame, bank: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    l = ledger.copy()
    b = bank.copy()

    l["ledger_id"] = l["ledger_id"].map(_clean_text)
    l["ledger_desc_norm"] = l["ledger_desc"].map(_norm_desc)
    l["ledger_dt"] = l["ledger_date"].map(_parse_date)
    l["ledger_amount_cents"] = l["ledger_amount_raw"].map(_parse_amount_cents)

    b["bank_id"] = b["bank_id"].map(_clean_text)
    b["bank_desc_norm"] = b["bank_desc"].map(_norm_desc)
    b["bank_dt"] = b["bank_date"].map(_parse_date)
    b["bank_amount_cents"] = b["bank_amount_raw"].map(_parse_amount_cents)

    # Treat missing IDs as invalid for auditability
    l["ledger_valid"] = (l["ledger_dt"].notna()) & (l["ledger_amount_cents"].notna()) & (l["ledger_id"] != "")
    b["bank_valid"] = (b["bank_dt"].notna()) & (b["bank_amount_cents"].notna()) & (b["bank_id"] != "")

    l["exact_key"] = (
        l["ledger_dt"].dt.strftime("%Y-%m-%d").fillna("")
        + "|"
        + l["ledger_amount_cents"].fillna(-999999).astype(int).astype(str)
        + "|"
        + l["ledger_desc_norm"].fillna("")
    )

    b["exact_key"] = (
        b["bank_dt"].dt.strftime("%Y-%m-%d").fillna("")
        + "|"
        + b["bank_amount_cents"].fillna(-999999).astype(int).astype(str)
        + "|"
        + b["bank_desc_norm"].fillna("")
    )

    return l, b


def _jaccard(a: str, b: str) -> float:
    sa = set([t for t in a.split() if t])
    sb = set([t for t in b.split() if t])
    if not sa and not sb:
        return 1.0
    if not sa or not sb:
        return 0.0
    inter = len(sa & sb)
    union = len(sa | sb)
    return inter / union if union else 0.0


def match_records(ledger: pd.DataFrame, bank: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    l = ledger.reset_index(drop=True).copy()
    b = bank.reset_index(drop=True).copy()

    used_bank = set()
    matched_ledger = set()
    matched_rows = []

    key_to_bank_idxs: dict[str, list[int]] = {}
    for i, row in b.iterrows():
        if not row["bank_valid"]:
            continue
        key_to_bank_idxs.setdefault(row["exact_key"], []).append(i)

    for k in key_to_bank_idxs:
        key_to_bank_idxs[k].sort(key=lambda idx: b.at[idx, "bank_id"])

    # Pass 1: exact key
    for _, lrow in l[l["ledger_valid"]].iterrows():
        k = lrow["exact_key"]
        candidates = key_to_bank_idxs.get(k, [])
        chosen = None
        for bi in candidates:
            if bi not in used_bank:
                chosen = bi
                break
        if chosen is None:
            continue

        used_bank.add(chosen)
        matched_ledger.add(lrow["ledger_id"])
        brow = b.loc[chosen]

        matched_rows.append(
            {
                "match_reason": "exact_key",
                "ledger_id": lrow["ledger_id"],
                "ledger_date": lrow["ledger_dt"].strftime("%Y-%m-%d"),
                "ledger_amount": lrow["ledger_amount_cents"] / 100.0,
                "ledger_desc": lrow["ledger_desc"],
                "bank_id": brow["bank_id"],
                "bank_date": brow["bank_dt"].strftime("%Y-%m-%d"),
                "bank_amount": brow["bank_amount_cents"] / 100.0,
                "bank_desc": brow["bank_desc"],
                "amount_diff": (lrow["ledger_amount_cents"] - brow["bank_amount_cents"]) / 100.0,
                "date_diff_days": int((lrow["ledger_dt"] - brow["bank_dt"]).days),
                "desc_similarity": _jaccard(lrow["ledger_desc_norm"], brow["bank_desc_norm"]),
            }
        )

    # Pass 2: fallback (amount within 2 cents, date within 2 days, description similarity)
    for _, lrow in l[l["ledger_valid"]].iterrows():
        if lrow["ledger_id"] in matched_ledger:
            continue

        l_cents = int(lrow["ledger_amount_cents"])
        l_dt = lrow["ledger_dt"]
        l_desc = lrow["ledger_desc_norm"]

        best = None  # (score, bank_index, reason, sim)
        for bi, brow in b[b["bank_valid"]].iterrows():
            if bi in used_bank:
                continue

            b_cents = int(brow["bank_amount_cents"])
            b_dt = brow["bank_dt"]
            b_desc = brow["bank_desc_norm"]

            amount_diff_cents = abs(l_cents - b_cents)
            if amount_diff_cents > 2:
                continue

            day_diff = abs(int((l_dt - b_dt).days))
            if day_diff > 2:
                continue

            sim = _jaccard(l_desc, b_desc)

            if l_desc and b_desc:
                if sim < 0.50:
                    continue
                reason = "fallback_amount_date_desc"
            else:
                if day_diff != 0 or amount_diff_cents != 0:
                    continue
                reason = "fallback_amount_date_missing_desc"

            score = (sim * 1000) - (day_diff * 10) - (amount_diff_cents * 2)

            if best is None:
                best = (score, bi, reason, sim)
            else:
                best_score, best_bi, _, _ = best
                if score > best_score or (score == best_score and brow["bank_id"] < b.loc[best_bi, "bank_id"]):
                    best = (score, bi, reason, sim)

        if best is None:
            continue

        _, chosen_bi, reason, sim = best
        used_bank.add(chosen_bi)
        matched_ledger.add(lrow["ledger_id"])
        brow = b.loc[chosen_bi]

        matched_rows.append(
            {
                "match_reason": reason,
                "ledger_id": lrow["ledger_id"],
                "ledger_date": lrow["ledger_dt"].strftime("%Y-%m-%d"),
                "ledger_amount": lrow["ledger_amount_cents"] / 100.0,
                "ledger_desc": lrow["ledger_desc"],
                "bank_id": brow["bank_id"],
                "bank_date": brow["bank_dt"].strftime("%Y-%m-%d"),
                "bank_amount": brow["bank_amount_cents"] / 100.0,
                "bank_desc": brow["bank_desc"],
                "amount_diff": (lrow["ledger_amount_cents"] - brow["bank_amount_cents"]) / 100.0,
                "date_diff_days": int((lrow["ledger_dt"] - brow["bank_dt"]).days),
                "desc_similarity": sim,
            }
        )

    matched_df = pd.DataFrame(matched_rows)

    # Exceptions
    exceptions = []

    for _, r in l[~l["ledger_valid"]].iterrows():
        reason_parts = []
        if _clean_text(r.get("ledger_id", "")) == "":
            reason_parts.append("missing_ledger_id")
        if pd.isna(r.get("ledger_dt", pd.NaT)):
            reason_parts.append("invalid_date")
        if r.get("ledger_amount_cents", None) is None:
            reason_parts.append("invalid_amount")
        exceptions.append(
            {
                "side": "ledger",
                "source_id": r.get("ledger_id", ""),
                "date": _clean_text(r.get("ledger_date", "")),
                "amount": _clean_text(r.get("ledger_amount_raw", "")),
                "description": r.get("ledger_desc", ""),
                "reason": "invalid_row:" + ",".join(reason_parts) if reason_parts else "invalid_row",
            }
        )

    matched_ledger_ids = set(matched_df["ledger_id"].tolist()) if not matched_df.empty else set()
    for _, r in l[l["ledger_valid"]].iterrows():
        if r["ledger_id"] not in matched_ledger_ids:
            exceptions.append(
                {
                    "side": "ledger",
                    "source_id": r["ledger_id"],
                    "date": r["ledger_dt"].strftime("%Y-%m-%d"),
                    "amount": r["ledger_amount_cents"] / 100.0,
                    "description": r["ledger_desc"],
                    "reason": "unmatched_ledger",
                }
            )

    for _, r in b[~b["bank_valid"]].iterrows():
        reason_parts = []
        if _clean_text(r.get("bank_id", "")) == "":
            reason_parts.append("missing_bank_id")
        if pd.isna(r.get("bank_dt", pd.NaT)):
            reason_parts.append("invalid_date")
        if r.get("bank_amount_cents", None) is None:
            reason_parts.append("invalid_amount")
        exceptions.append(
            {
                "side": "bank",
                "source_id": r.get("bank_id", ""),
                "date": _clean_text(r.get("bank_date", "")),
                "amount": _clean_text(r.get("bank_amount_raw", "")),
                "description": r.get("bank_desc", ""),
                "reason": "invalid_row:" + ",".join(reason_parts) if reason_parts else "invalid_row",
            }
        )

    matched_bank_ids = set(matched_df["bank_id"].tolist()) if not matched_df.empty else set()
    for _, r in b[b["bank_valid"]].iterrows():
        if r["bank_id"] not in matched_bank_ids:
            exceptions.append(
                {
                    "side": "bank",
                    "source_id": r["bank_id"],
                    "date": r["bank_dt"].strftime("%Y-%m-%d"),
                    "amount": r["bank_amount_cents"] / 100.0,
                    "description": r["bank_desc"],
                    "reason": "unmatched_bank",
                }
            )

    exceptions_df = pd.DataFrame(exceptions)
    return matched_df, exceptions_df


def build_summary(ledger: pd.DataFrame, bank: pd.DataFrame, matched_df: pd.DataFrame, exceptions_df: pd.DataFrame) -> dict:
    ledger_valid = ledger[ledger["ledger_valid"]]
    bank_valid = bank[bank["bank_valid"]]

    def cents_sum(series) -> int:
        vals = [v for v in series.tolist() if v is not None]
        return int(sum(vals)) if vals else 0

    ledger_total_cents = cents_sum(ledger_valid["ledger_amount_cents"])
    bank_total_cents = cents_sum(bank_valid["bank_amount_cents"])

    matched_total_ledger = float(matched_df["ledger_amount"].sum()) if not matched_df.empty else 0.0
    matched_total_bank = float(matched_df["bank_amount"].sum()) if not matched_df.empty else 0.0

    reason_breakdown = matched_df["match_reason"].value_counts().to_dict() if not matched_df.empty else {}
    exception_breakdown = exceptions_df["reason"].value_counts().to_dict() if not exceptions_df.empty else {}

    ledger_dup_keys = int(ledger_valid["exact_key"].duplicated().sum()) if not ledger_valid.empty else 0
    bank_dup_keys = int(bank_valid["exact_key"].duplicated().sum()) if not bank_valid.empty else 0

    return {
        "run_date": date.today().isoformat(),
        "counts": {
            "ledger_rows": int(len(ledger)),
            "bank_rows": int(len(bank)),
            "ledger_valid_rows": int(len(ledger_valid)),
            "bank_valid_rows": int(len(bank_valid)),
            "matched_rows": int(len(matched_df)),
            "exception_rows": int(len(exceptions_df)),
        },
        "totals": {
            "ledger_total": round(ledger_total_cents / 100.0, 2),
            "bank_total": round(bank_total_cents / 100.0, 2),
            "matched_ledger_total": round(matched_total_ledger, 2),
            "matched_bank_total": round(matched_total_bank, 2),
        },
        "breakdowns": {
            "match_reason_counts": reason_breakdown,
            "exception_reason_counts": exception_breakdown,
            "duplicate_exact_keys": {"ledger": ledger_dup_keys, "bank": bank_dup_keys},
        },
    }


def _autosize_columns(ws, max_width: int = 50):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        best = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            best = max(best, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_width, max(10, best + 2))


def _write_df_sheet(wb: Workbook, title: str, df: pd.DataFrame, table_name: str):
    ws = wb.create_sheet(title=title)

    if df is None or df.empty:
        ws["A1"] = "No rows"
        ws["A1"].font = Font(bold=True)
        return

    # Header
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Rows
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    ws.freeze_panes = "A2"

    max_row = ws.max_row
    max_col = ws.max_column
    ref = f"A1:{get_column_letter(max_col)}{max_row}"

    tab = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    _autosize_columns(ws)


def write_excel_report(outdir: str, summary: dict, matched_df: pd.DataFrame, exceptions_df: pd.DataFrame) -> Path:
    out = Path(outdir)
    out.mkdir(parents=True, exist_ok=True)
    xlsx_path = out / "recon_report.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Finance Ops Reconciliation Summary (Synthetic Data)"
    ws["A1"].font = Font(bold=True, size=14)

    row = 3
    ws[f"A{row}"] = "Run date"
    ws[f"B{row}"] = summary.get("run_date", "")
    ws[f"A{row}"].font = Font(bold=True)
    row += 2

    ws[f"A{row}"] = "Counts"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    for k, v in summary["counts"].items():
        ws[f"A{row}"] = k
        ws[f"B{row}"] = v
        row += 1

    row += 1
    ws[f"A{row}"] = "Totals"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    for k, v in summary["totals"].items():
        ws[f"A{row}"] = k
        ws[f"B{row}"] = v
        row += 1

    row += 1
    ws[f"A{row}"] = "Breakdowns"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    ws[f"A{row}"] = "match_reason_counts"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    for k, v in summary["breakdowns"]["match_reason_counts"].items():
        ws[f"A{row}"] = k
        ws[f"B{row}"] = v
        row += 1

    row += 1
    ws[f"A{row}"] = "exception_reason_counts"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    for k, v in summary["breakdowns"]["exception_reason_counts"].items():
        ws[f"A{row}"] = k
        ws[f"B{row}"] = v
        row += 1

    row += 1
    ws[f"A{row}"] = "duplicate_exact_keys"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    for k, v in summary["breakdowns"]["duplicate_exact_keys"].items():
        ws[f"A{row}"] = k
        ws[f"B{row}"] = v
        row += 1

    _autosize_columns(ws)

    _write_df_sheet(wb, "Exceptions", exceptions_df, "ExceptionsTable")
    _write_df_sheet(wb, "Matched", matched_df, "MatchedTable")

    wb.save(xlsx_path)
    return xlsx_path


def write_outputs(outdir: str, matched_df: pd.DataFrame, exceptions_df: pd.DataFrame, summary: dict) -> None:
    out = Path(outdir)
    out.mkdir(parents=True, exist_ok=True)

    matched_path = out / "matched.csv"
    exceptions_path = out / "exceptions.csv"
    summary_path = out / "summary.json"

    matched_df.to_csv(matched_path, index=False)
    exceptions_df.to_csv(exceptions_path, index=False)
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    xlsx_path = write_excel_report(outdir, summary, matched_df, exceptions_df)

    print(f"WROTE: {matched_path}")
    print(f"WROTE: {exceptions_path}")
    print(f"WROTE: {summary_path}")
    print(f"WROTE: {xlsx_path}")


def main():
    p = argparse.ArgumentParser(description="Reconcile ledger CSV vs bank CSV (synthetic demo).")
    p.add_argument("--ledger", required=True, help="Path to ledger_export.csv")
    p.add_argument("--bank", required=True, help="Path to bank_export.csv")
    p.add_argument("--outdir", required=True, help="Output directory")
    args = p.parse_args()

    ledger_raw, bank_raw = load_data(args.ledger, args.bank)
    ledger_std, bank_std = standardize_columns(ledger_raw, bank_raw)
    ledger_clean, bank_clean = clean_fields(ledger_std, bank_std)

    matched_df, exceptions_df = match_records(ledger_clean, bank_clean)
    summary = build_summary(ledger_clean, bank_clean, matched_df, exceptions_df)
    write_outputs(args.outdir, matched_df, exceptions_df, summary)

    c = summary["counts"]
    t = summary["totals"]
    print("SUMMARY:")
    print(f"  matched_rows={c['matched_rows']} exceptions_rows={c['exception_rows']}")
    print(f"  ledger_total={t['ledger_total']} bank_total={t['bank_total']}")
    print(f"  matched_ledger_total={t['matched_ledger_total']} matched_bank_total={t['matched_bank_total']}")


if __name__ == "__main__":
    main()
